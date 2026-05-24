import base64
import binascii
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import io
import re
from typing import Callable
import unicodedata
from zipfile import BadZipFile

from flask import Blueprint, g, jsonify, request
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from psycopg2.extras import DictCursor, Json, execute_values

MAX_ARCHIVO_BYTES = 2 * 1024 * 1024
ALLOWED_EXTRACTO_ESTADOS = {
    "Cargado",
    "Validado",
    "Parcialmente Consolidado",
    "Consolidado",
    "Identificado",
}
BASE_CONSOLIDADA_ROLE = {"gestor_cuenta_bancaria"}
CUENTA_CONSOLIDADA_ROLES = {"gestor_cuenta_bancaria", "gestor"}
ACTIVIDAD_TIPOS = {
    "Asignacion de Gestor",
    "Agregar Adjuntos",
    "Observaciones",
    "Confirmacion",
    "Cambio de Estado",
}
ALLOWED_CONSOLIDADA_ESTADOS = {"Pendiente", "En Progreso", "Concluido"}
EXPECTED_XLSX_HEADERS = {
    "FECHA": "fecha",
    "OFICINA": "oficina",
    "NUMERO DE DOCUMENTO": "numero_de_documento",
    "DESCRIPCION": "descripcion",
    "DEBITO": "debito",
    "CREDITO": "credito",
    "SALDO": "saldo",
}


def _decode_archivo_blob(raw_value: object):
    if not isinstance(raw_value, str) or not raw_value.strip():
        return None, "archivo is required and must be a base64 string"

    try:
        decoded = base64.b64decode(raw_value, validate=True)
    except (binascii.Error, ValueError):
        return None, "archivo must be valid base64"

    if len(decoded) > MAX_ARCHIVO_BYTES:
        return None, "archivo exceeds 2MB limit"

    return decoded, None


def _normalize_header(value: object) -> str:
    if value is None:
        return ""

    normalized = unicodedata.normalize("NFKD", str(value).strip())
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.upper()


def _to_fecha_string(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    return str(value).strip()[:16]


def _normalize_fecha_for_db(value: object) -> str:
    raw = _to_fecha_string(value)
    if not raw:
        return ""

    candidate = raw[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(candidate, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _to_decimal(value: object):
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    cleaned = str(value).strip().replace(" ", "")
    if not cleaned:
        return None

    if "." in cleaned and "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")

    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def _cell_value(row_values: tuple[object, ...], col_idx: int):
    if col_idx >= len(row_values):
        return None
    return row_values[col_idx]


def _extract_movimientos_from_xlsx(archivo_bytes: bytes, validate_rows: bool = True):
    try:
        workbook = load_workbook(io.BytesIO(archivo_bytes), data_only=True, read_only=True)
    except (BadZipFile, InvalidFileException, ValueError) as exc:
        raise ValueError("archivo is not a valid XLSX file") from exc

    sheet = workbook.active
    if sheet is None:
        workbook.close()
        raise ValueError("XLSX file does not contain an active worksheet")

    header_row_idx = None
    header_map: dict[str, int] = {}

    for row_idx, row_values in enumerate(sheet.iter_rows(values_only=True), start=1):
        normalized_headers = [_normalize_header(cell) for cell in row_values]
        temp_map: dict[str, int] = {}

        for col_idx, header in enumerate(normalized_headers):
            canonical = EXPECTED_XLSX_HEADERS.get(header)
            if canonical:
                temp_map[canonical] = col_idx

        if len(temp_map) == len(EXPECTED_XLSX_HEADERS):
            header_row_idx = row_idx
            header_map = temp_map
            break

    if header_row_idx is None:
        raise ValueError(
            "XLSX header not found. Expected columns: FECHA, OFICINA, NUMERO DE DOCUMENTO, "
            "DESCRIPCION, DEBITO, CREDITO, SALDO"
        )

    movimientos = []
    row_errors = []
    try:
        for row_number, row_values in enumerate(
            sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            raw_fecha = _cell_value(row_values, header_map["fecha"])
            raw_oficina = _cell_value(row_values, header_map["oficina"])
            raw_numero_documento = _cell_value(row_values, header_map["numero_de_documento"])
            raw_descripcion = _cell_value(row_values, header_map["descripcion"])
            raw_debito = _cell_value(row_values, header_map["debito"])
            raw_credito = _cell_value(row_values, header_map["credito"])
            raw_saldo = _cell_value(row_values, header_map["saldo"])

            fecha = _normalize_fecha_for_db(raw_fecha)
            oficina = str(raw_oficina or "").strip() or None
            numero_de_documento = str(raw_numero_documento or "").strip() or None
            descripcion = str(raw_descripcion or "").strip() or None
            debito = _to_decimal(raw_debito)
            credito = _to_decimal(raw_credito)
            saldo = _to_decimal(raw_saldo)

            is_empty = not any([
                fecha,
                oficina,
                numero_de_documento,
                descripcion,
                debito is not None,
                credito is not None,
                saldo is not None,
            ])
            if is_empty:
                continue

            current_row_errors = []
            if validate_rows:
                if not fecha:
                    current_row_errors.append(f"row {row_number}: FECHA is required")

                if not numero_de_documento:
                    current_row_errors.append(f"row {row_number}: NUMERO DE DOCUMENTO is required")

                if not _is_blank(raw_debito) and debito is None:
                    current_row_errors.append(f"row {row_number}: DEBITO must be a valid number")

                if not _is_blank(raw_credito) and credito is None:
                    current_row_errors.append(f"row {row_number}: CREDITO must be a valid number")

                if not _is_blank(raw_saldo) and saldo is None:
                    current_row_errors.append(f"row {row_number}: SALDO must be a valid number")

                if current_row_errors:
                    row_errors.extend(current_row_errors)

            error_validacion = "; ".join(current_row_errors) if current_row_errors else None
            estado_validacion = "Error" if current_row_errors else ("Validado" if validate_rows else "Pendiente")

            movimientos.append(
                {
                    "fila_excel": row_number,
                    "fecha": fecha,
                    "oficina": oficina,
                    "numero_de_documento": numero_de_documento,
                    "descripcion": descripcion,
                    "debito": debito,
                    "credito": credito,
                    "saldo": saldo,
                    "estado_validacion": estado_validacion,
                    "error_validacion": error_validacion,
                }
            )
    finally:
        workbook.close()

    return movimientos, row_errors


def _error_response(status: int, code: str, message: str, details: dict | None = None):
    return (
        jsonify(
            {
                "error": {
                    "code": code,
                    "message": message,
                    "details": details or {},
                }
            }
        ),
        status,
    )


def create_bank_accounts_blueprint(
    get_db_connection: Callable,
    token_required: Callable,
) -> Blueprint:
    bank_accounts_bp = Blueprint("bank_accounts", __name__)

    @bank_accounts_bp.route("/cuentas", methods=["GET"])
    @token_required()
    def list_cuentas():
        query = """
            SELECT id, nombre_cuenta, numero_cuenta, entidad_bancaria, estado
            FROM cuentas
            ORDER BY id ASC;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(query)
                rows = cur.fetchall()

        return jsonify(
            {
                "items": [
                    {
                        "id": row["id"],
                        "nombre_cuenta": row["nombre_cuenta"],
                        "numero_cuenta": row["numero_cuenta"],
                        "entidad_bancaria": row["entidad_bancaria"],
                        "estado": row["estado"],
                    }
                    for row in rows
                ]
            }
        )

    @bank_accounts_bp.route("/cuentas/<int:cuenta_id>/", methods=["GET"])
    @token_required()
    def get_cuenta_with_extractos(cuenta_id: int):
        cuenta_query = """
            SELECT id, nombre_cuenta, numero_cuenta, entidad_bancaria, estado
            FROM cuentas
            WHERE id = %s;
        """

        extractos_query = """
            SELECT id, nombre_del_archivo, fecha, tipo, fecha_carga, estado, errores_validacion
            FROM extracto_bancario
            WHERE cuenta_id = %s
            ORDER BY fecha_carga DESC, id DESC;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(cuenta_query, (cuenta_id,))
                cuenta = cur.fetchone()
                if cuenta is None:
                    return jsonify({"error": "Cuenta not found"}), 404

                cur.execute(extractos_query, (cuenta_id,))
                extractos = cur.fetchall()

        return jsonify(
            {
                "cuenta": {
                    "id": cuenta["id"],
                    "nombre_cuenta": cuenta["nombre_cuenta"],
                    "numero_cuenta": cuenta["numero_cuenta"],
                    "entidad_bancaria": cuenta["entidad_bancaria"],
                    "estado": cuenta["estado"],
                },
                "extractos_bancarios": [
                    {
                        "id": row["id"],
                        "nombre_del_archivo": row["nombre_del_archivo"],
                        "fecha": row["fecha"],
                        "tipo": row["tipo"],
                        "fecha_carga": row["fecha_carga"].isoformat() if row["fecha_carga"] else None,
                        "estado": row["estado"],
                        "errores_validacion": row["errores_validacion"],
                    }
                    for row in extractos
                ],
            }
        )

    @bank_accounts_bp.route("/cuentas/<int:cuenta_id>/extracto-bancario", methods=["POST"])
    @token_required()
    def create_extracto_bancario(cuenta_id: int):
        payload = request.get_json(silent=True) or {}
        nombre_del_archivo = (payload.get("nombre_del_archivo") or "").strip()
        fecha = (payload.get("fecha") or "").strip()
        tipo = (payload.get("tipo") or "XLSX").strip().upper()
        estado = (payload.get("estado") or "Cargado").strip()

        if not nombre_del_archivo:
            return jsonify({"error": "nombre_del_archivo is required"}), 400
        if not fecha:
            return jsonify({"error": "fecha is required"}), 400
        if len(fecha) > 16:
            return jsonify({"error": "fecha must have at most 16 characters"}), 400
        if tipo != "XLSX":
            return jsonify({"error": "tipo must be XLSX"}), 400
        if estado not in ALLOWED_EXTRACTO_ESTADOS:
            return jsonify({"error": "estado is invalid"}), 400

        archivo, decode_error = _decode_archivo_blob(payload.get("archivo"))
        if decode_error:
            return jsonify({"error": decode_error}), 400
        if archivo is None:
            return jsonify({"error": "archivo is required"}), 400

        estado_persistido = "Cargado"
        errores_validacion = None

        cuenta_exists_query = "SELECT id FROM cuentas WHERE id = %s;"
        insert_extracto_query = """
            INSERT INTO extracto_bancario (
                cuenta_id,
                nombre_del_archivo,
                archivo,
                fecha,
                tipo,
                estado,
                errores_validacion
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id, fecha_carga;
        """
        insert_movimientos_query = """
            INSERT INTO movimientos_bancarios (
                extracto_bancario_id,
                fila_excel,
                fecha,
                oficina,
                numero_de_documento,
                descripcion,
                debito,
                credito,
                saldo,
                estado_validacion,
                error_validacion
            )
            VALUES %s;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(cuenta_exists_query, (cuenta_id,))
                cuenta = cur.fetchone()
                if cuenta is None:
                    return jsonify({"error": "Cuenta not found"}), 404

                try:
                    movimientos, _ = _extract_movimientos_from_xlsx(archivo, validate_rows=False)
                except ValueError as exc:
                    return jsonify({"error": str(exc)}), 400

                cur.execute(
                    insert_extracto_query,
                    (
                        cuenta_id,
                        nombre_del_archivo,
                        archivo,
                        fecha,
                        tipo,
                        estado_persistido,
                        errores_validacion,
                    ),
                )
                created = cur.fetchone()

                movimiento_rows = [
                    (
                        created["id"],
                        movimiento["fila_excel"],
                        movimiento["fecha"],
                        movimiento["oficina"],
                        movimiento["numero_de_documento"],
                        movimiento["descripcion"],
                        movimiento["debito"],
                        movimiento["credito"],
                        movimiento["saldo"],
                        movimiento["estado_validacion"],
                        movimiento["error_validacion"],
                    )
                    for movimiento in movimientos
                ]

                if movimiento_rows:
                    execute_values(cur, insert_movimientos_query, movimiento_rows)

                conn.commit()

        return (
            jsonify(
                {
                    "id": created["id"],
                    "cuenta_id": cuenta_id,
                    "nombre_del_archivo": nombre_del_archivo,
                    "fecha": fecha,
                    "tipo": tipo,
                    "estado": estado_persistido,
                    "fecha_carga": created["fecha_carga"].isoformat() if created["fecha_carga"] else None,
                    "movimientos_creados": len(movimiento_rows),
                    "errores_validacion": [],
                }
            ),
            201,
        )

    @bank_accounts_bp.route("/cuentas/<int:cuenta_id>/extracto-bancario/<int:extracto_id>/validar", methods=["POST"])
    @token_required()
    def validar_extracto_bancario(cuenta_id: int, extracto_id: int):
        extracto_query = """
            SELECT id, cuenta_id, archivo
            FROM extracto_bancario
            WHERE id = %s AND cuenta_id = %s;
        """
        update_extracto_query = """
            UPDATE extracto_bancario
            SET estado = %s, errores_validacion = %s
            WHERE id = %s;
        """
        insert_movimientos_query = """
            INSERT INTO movimientos_bancarios (
                extracto_bancario_id,
                fila_excel,
                fecha,
                oficina,
                numero_de_documento,
                descripcion,
                debito,
                credito,
                saldo,
                estado_validacion,
                error_validacion
            )
            VALUES %s;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(extracto_query, (extracto_id, cuenta_id))
                extracto = cur.fetchone()
                if extracto is None:
                    return jsonify({"error": "Extracto bancario not found"}), 404

                try:
                    movimientos, validation_errors = _extract_movimientos_from_xlsx(extracto["archivo"])
                except ValueError as exc:
                    cur.execute(update_extracto_query, ("Cargado", str(exc), extracto_id))
                    conn.commit()
                    return jsonify(
                        {
                            "extracto_id": extracto_id,
                            "cuenta_id": cuenta_id,
                            "estado": "Cargado",
                            "movimientos_creados": 0,
                            "errores_validacion": [str(exc)],
                        }
                    )

                if validation_errors:
                    max_errors = 20
                    shown_errors = validation_errors[:max_errors]
                    suffix = (
                        ""
                        if len(validation_errors) <= max_errors
                        else f" (+{len(validation_errors) - max_errors} more)"
                    )
                    errores_extracto = "Invalid XLSX data: " + "; ".join(shown_errors) + suffix
                    estado_extracto = "Cargado"
                else:
                    errores_extracto = None
                    estado_extracto = "Validado"

                cur.execute(
                    "DELETE FROM movimientos_bancarios WHERE extracto_bancario_id = %s;",
                    (extracto_id,),
                )

                movimiento_rows = [
                    (
                        extracto_id,
                        movimiento["fila_excel"],
                        movimiento["fecha"],
                        movimiento["oficina"],
                        movimiento["numero_de_documento"],
                        movimiento["descripcion"],
                        movimiento["debito"],
                        movimiento["credito"],
                        movimiento["saldo"],
                        movimiento["estado_validacion"],
                        movimiento["error_validacion"],
                    )
                    for movimiento in movimientos
                ]

                if movimiento_rows:
                    execute_values(cur, insert_movimientos_query, movimiento_rows)

                cur.execute(update_extracto_query, (estado_extracto, errores_extracto, extracto_id))
                conn.commit()

        return jsonify(
            {
                "extracto_id": extracto_id,
                "cuenta_id": cuenta_id,
                "estado": estado_extracto,
                "movimientos_creados": len(movimiento_rows),
                "errores_validacion": validation_errors,
            }
        )

    @bank_accounts_bp.route("/cuentas/<int:cuenta_id>/extracto-bancario/<int:extracto_id>", methods=["GET"])
    @token_required()
    def get_movimientos_bancarios(cuenta_id: int, extracto_id: int):
        extracto_query = """
            SELECT id, cuenta_id, nombre_del_archivo, fecha, tipo, fecha_carga, estado, errores_validacion
            FROM extracto_bancario
            WHERE id = %s AND cuenta_id = %s;
        """

        movimientos_query = """
            SELECT
                id,
                extracto_bancario_id,
                fila_excel,
                fecha,
                oficina,
                numero_de_documento,
                descripcion,
                debito,
                credito,
                saldo,
                estado_validacion,
                error_validacion,
                error_carga
            FROM movimientos_bancarios
            WHERE extracto_bancario_id = %s
            ORDER BY fila_excel ASC, id ASC;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(extracto_query, (extracto_id, cuenta_id))
                extracto = cur.fetchone()
                if extracto is None:
                    return jsonify({"error": "Extracto bancario not found"}), 404

                cur.execute(movimientos_query, (extracto_id,))
                movimientos = cur.fetchall()

        return jsonify(
            {
                "cuenta_id": cuenta_id,
                "extracto_bancario": {
                    "id": extracto["id"],
                    "nombre_del_archivo": extracto["nombre_del_archivo"],
                    "fecha": extracto["fecha"],
                    "tipo": extracto["tipo"],
                    "fecha_carga": extracto["fecha_carga"].isoformat() if extracto["fecha_carga"] else None,
                    "estado": extracto["estado"],
                    "errores_validacion": extracto["errores_validacion"],
                },
                "movimientos": [
                    {
                        "id": row["id"],
                        "extracto_bancario_id": row["extracto_bancario_id"],
                        "fila_excel": row["fila_excel"],
                        "fecha": row["fecha"],
                        "oficina": row["oficina"],
                        "numero_de_documento": row["numero_de_documento"],
                        "descripcion": row["descripcion"],
                        "debito": float(row["debito"]) if row["debito"] is not None else None,
                        "credito": float(row["credito"]) if row["credito"] is not None else None,
                        "saldo": float(row["saldo"]) if row["saldo"] is not None else None,
                        "estado_validacion": row["estado_validacion"],
                        "error_validacion": row["error_validacion"],
                        "error_carga": row["error_carga"],
                    }
                    for row in movimientos
                ],
            }
        )

    @bank_accounts_bp.route(
        "/cuentas/<int:cuenta_id>/extracto-bancario/<int:extracto_id>/cargar",
        methods=["POST"],
    )
    @token_required(required_roles=BASE_CONSOLIDADA_ROLE)
    def cargar_extracto_a_base_consolidada(cuenta_id: int, extracto_id: int):
        cuenta_query = "SELECT id FROM cuentas WHERE id = %s;"
        extracto_query = """
            SELECT id, cuenta_id, estado, errores_validacion
            FROM extracto_bancario
            WHERE id = %s AND cuenta_id = %s;
        """
        movimientos_invalidos_query = """
            SELECT COUNT(*) AS total
            FROM movimientos_bancarios
            WHERE extracto_bancario_id = %s
              AND (
                  estado_validacion = 'Error'
                  OR error_validacion IS NOT NULL
                  OR numero_de_documento IS NULL
                  OR btrim(numero_de_documento) = ''
              );
        """
        movimientos_query = """
            SELECT id, fecha, numero_de_documento, descripcion, credito, saldo,
                   estado_validacion, error_validacion
            FROM movimientos_bancarios
            WHERE extracto_bancario_id = %s
            ORDER BY fila_excel ASC, id ASC;
        """
        reset_errores_query = """
            UPDATE movimientos_bancarios
            SET error_carga = NULL
            WHERE extracto_bancario_id = %s;
        """
        mark_error_carga_query = """
            UPDATE movimientos_bancarios
            SET error_carga = %s
            WHERE id = %s;
        """
        insert_consolidada_query = """
            INSERT INTO cuenta_consolidada (
                cuenta_id,
                extracto_bancario_id,
                movimiento_bancario_id,
                fecha_movimiento,
                numero_de_documento,
                descripcion,
                monto_pago,
                saldo,
                estado
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente')
            ON CONFLICT (cuenta_id, numero_de_documento, fecha_movimiento) DO NOTHING
            RETURNING id;
        """
        update_extracto_estado_query = """
            UPDATE extracto_bancario
            SET estado = %s,
                resumen_carga = %s
            WHERE id = %s;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(cuenta_query, (cuenta_id,))
                if cur.fetchone() is None:
                    return _error_response(
                        404,
                        "CUENTA_NO_ENCONTRADA",
                        "La cuenta indicada no existe.",
                        {"cuenta_id": cuenta_id},
                    )

                cur.execute(extracto_query, (extracto_id, cuenta_id))
                extracto = cur.fetchone()
                if extracto is None:
                    return _error_response(
                        404,
                        "EXTRACTO_NO_ENCONTRADO",
                        "El extracto bancario no existe para la cuenta indicada.",
                        {"cuenta_id": cuenta_id, "extracto_id": extracto_id},
                    )

                if extracto["estado"] != "Validado":
                    return _error_response(
                        409,
                        "EXTRACTO_NO_VALIDADO",
                        "El extracto bancario debe estar en estado Validado para cargarse a Base Consolidada.",
                        {
                            "cuenta_id": cuenta_id,
                            "extracto_id": extracto_id,
                            "estado_actual": extracto["estado"],
                        },
                    )

                if extracto["errores_validacion"]:
                    return _error_response(
                        409,
                        "EXTRACTO_CON_ERRORES_VALIDACION",
                        "El extracto contiene errores de validacion. Corrija los datos y cargue nuevamente el extracto.",
                        {"cuenta_id": cuenta_id, "extracto_id": extracto_id},
                    )

                cur.execute(movimientos_invalidos_query, (extracto_id,))
                invalidos = cur.fetchone()["total"]
                if invalidos > 0:
                    return _error_response(
                        409,
                        "EXTRACTO_CON_ERRORES_VALIDACION",
                        "El extracto contiene errores de validacion. Corrija los datos y cargue nuevamente el extracto.",
                        {
                            "cuenta_id": cuenta_id,
                            "extracto_id": extracto_id,
                            "filas_con_error": invalidos,
                        },
                    )

                cur.execute(movimientos_query, (extracto_id,))
                movimientos = cur.fetchall()
                if not movimientos:
                    return _error_response(
                        422,
                        "EXTRACTO_SIN_MOVIMIENTOS_ELEGIBLES",
                        "El extracto no contiene movimientos de pago elegibles para Base Consolidada.",
                        {"cuenta_id": cuenta_id, "extracto_id": extracto_id, "regla": "credito > 0"},
                    )

                # Clear previous processing markers for a fresh load summary.
                cur.execute(reset_errores_query, (extracto_id,))

                consolidados_creados = 0
                consolidados_omitidos = 0
                consolidados_con_error = 0
                elegibles = 0

                for movimiento in movimientos:
                    credito = movimiento["credito"]
                    if credito is None or credito <= 0:
                        continue
                    elegibles += 1

                    if movimiento["estado_validacion"] == "Error" or movimiento["error_validacion"]:
                        consolidados_con_error += 1
                        cur.execute(
                            mark_error_carga_query,
                            ("No se puede cargar: fila con error de validacion.", movimiento["id"]),
                        )
                        continue

                    numero_de_documento = (movimiento["numero_de_documento"] or "").strip()
                    if not numero_de_documento:
                        consolidados_con_error += 1
                        cur.execute(
                            mark_error_carga_query,
                            ("No se puede cargar: NUMERO DE DOCUMENTO vacio.", movimiento["id"]),
                        )
                        continue

                    fecha_movimiento = _normalize_fecha_for_db(movimiento["fecha"])
                    if not fecha_movimiento:
                        consolidados_con_error += 1
                        cur.execute(
                            mark_error_carga_query,
                            ("No se puede cargar: FECHA invalida para normalizar a YYYY-MM-DD.", movimiento["id"]),
                        )
                        continue

                    cur.execute(
                        insert_consolidada_query,
                        (
                            cuenta_id,
                            extracto_id,
                            movimiento["id"],
                            fecha_movimiento,
                            numero_de_documento,
                            movimiento["descripcion"],
                            credito,
                            movimiento["saldo"],
                        ),
                    )
                    inserted = cur.fetchone()
                    if inserted:
                        consolidados_creados += 1
                    else:
                        cur.execute(
                            mark_error_carga_query,
                            (
                                "Registro omitido por idempotencia: ya existe Cuenta + Numero de Documento + Fecha.",
                                movimiento["id"],
                            ),
                        )
                        consolidados_omitidos += 1
                        consolidados_con_error += 1

                if elegibles == 0:
                    conn.rollback()
                    return _error_response(
                        422,
                        "EXTRACTO_SIN_MOVIMIENTOS_ELEGIBLES",
                        "El extracto no contiene movimientos de pago elegibles para Base Consolidada.",
                        {"cuenta_id": cuenta_id, "extracto_id": extracto_id, "regla": "credito > 0"},
                    )

                if consolidados_con_error > 0:
                    nuevo_estado = "Parcialmente Consolidado"
                else:
                    nuevo_estado = "Consolidado"

                resumen_carga = {
                    "cuenta_id": cuenta_id,
                    "extracto_id": extracto_id,
                    "consolidados_creados": consolidados_creados,
                    "consolidados_omitidos": consolidados_omitidos,
                    "consolidados_con_error": consolidados_con_error,
                    "estado_extracto": nuevo_estado,
                    "procesado_en": datetime.utcnow().isoformat() + "Z",
                }

                cur.execute(update_extracto_estado_query, (nuevo_estado, Json(resumen_carga), extracto_id))
                conn.commit()

        return jsonify(
            {
                "cuenta_id": cuenta_id,
                "extracto_id": extracto_id,
                "consolidados_creados": consolidados_creados,
                "consolidados_omitidos": consolidados_omitidos,
                "consolidados_con_error": consolidados_con_error,
                "estado_extracto": nuevo_estado,
                "resumen_carga": resumen_carga,
            }
        )

    @bank_accounts_bp.route("/cuenta-consolidada", methods=["GET"])
    @token_required(required_roles=CUENTA_CONSOLIDADA_ROLES)
    def list_cuenta_consolidada():
        params = []
        where = []

        def add_filter(field: str, value: str | None):
            if value is None or str(value).strip() == "":
                return
            params.append(value)
            where.append(f"cc.{field} = %s")

        add_filter("cuenta_id", request.args.get("cuenta_id"))
        add_filter("extracto_bancario_id", request.args.get("extracto_bancario_id"))
        add_filter("estado", request.args.get("estado"))
        add_filter("cliente_id", request.args.get("cliente_id"))
        add_filter("obligacion_id", request.args.get("obligacion_id"))
        add_filter("gestor_user_id", request.args.get("gestor_user_id"))

        where_sql = ""
        if where:
            where_sql = "WHERE " + " AND ".join(where)

        query = f"""
            SELECT
                cc.id,
                cc.cuenta_id,
                cc.extracto_bancario_id,
                cc.movimiento_bancario_id,
                cc.fecha_movimiento,
                cc.numero_de_documento,
                cc.descripcion,
                cc.monto_pago,
                cc.saldo,
                cc.cliente_id,
                cc.obligacion_id,
                cc.gestor_user_id,
                cc.observaciones,
                cc.estado,
                (
                    SELECT COUNT(*)
                    FROM cuenta_consolidada_adjuntos cca
                    WHERE cca.cuenta_consolidada_id = cc.id
                ) AS adjuntos_count
            FROM cuenta_consolidada cc
            {where_sql}
            ORDER BY cc.id DESC;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(query, tuple(params))
                rows = cur.fetchall()

        return jsonify(
            {
                "items": [
                    {
                        "id": row["id"],
                        "cuenta_id": row["cuenta_id"],
                        "extracto_bancario_id": row["extracto_bancario_id"],
                        "movimiento_bancario_id": row["movimiento_bancario_id"],
                        "fecha_movimiento": row["fecha_movimiento"],
                        "numero_de_documento": row["numero_de_documento"],
                        "descripcion": row["descripcion"],
                        "monto_pago": float(row["monto_pago"]) if row["monto_pago"] is not None else None,
                        "saldo": float(row["saldo"]) if row["saldo"] is not None else None,
                        "cliente_id": row["cliente_id"],
                        "obligacion_id": row["obligacion_id"],
                        "gestor_user_id": row["gestor_user_id"],
                        "observaciones": row["observaciones"],
                        "estado": row["estado"],
                        "adjuntos_count": row["adjuntos_count"],
                    }
                    for row in rows
                ]
            }
        )

    @bank_accounts_bp.route("/cuenta-consolidada/<int:consolidada_id>", methods=["PATCH"])
    @token_required(required_roles=CUENTA_CONSOLIDADA_ROLES)
    def patch_cuenta_consolidada(consolidada_id: int):
        payload = request.get_json(silent=True) or {}
        allowed_fields = {
            "cliente_id": payload.get("cliente_id"),
            "obligacion_id": payload.get("obligacion_id"),
            "gestor_user_id": payload.get("gestor_user_id"),
            "observaciones": payload.get("observaciones"),
            "estado": payload.get("estado"),
        }

        updates = []
        params = []
        for field, value in allowed_fields.items():
            if field not in payload:
                continue
            if field == "estado" and value not in ALLOWED_CONSOLIDADA_ESTADOS:
                return _error_response(400, "ESTADO_INVALIDO", "estado is invalid")
            updates.append(f"{field} = %s")
            params.append(value)

        if not updates:
            return _error_response(400, "SIN_CAMBIOS", "No updatable fields provided")

        params.append(consolidada_id)
        query = f"""
            UPDATE cuenta_consolidada
            SET {', '.join(updates)}, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id, estado;
        """

        select_estado_actual_query = """
            SELECT id, estado
            FROM cuenta_consolidada
            WHERE id = %s;
        """

        insert_actividad_query = """
            INSERT INTO cuenta_consolidada_actividad (
                cuenta_consolidada_id,
                tipo,
                fecha,
                comentario,
                usuario_id
            )
            VALUES (%s, 'Cambio de Estado', CURRENT_TIMESTAMP, %s, %s);
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(select_estado_actual_query, (consolidada_id,))
                consolidado_actual = cur.fetchone()
                if consolidado_actual is None:
                    return _error_response(404, "CONSOLIDADO_NO_ENCONTRADO", "Registro consolidado no encontrado")

                estado_anterior = consolidado_actual["estado"]
                cur.execute(query, tuple(params))
                row = cur.fetchone()
                if row is None:
                    return _error_response(404, "CONSOLIDADO_NO_ENCONTRADO", "Registro consolidado no encontrado")

                if "estado" in payload and payload.get("estado") != estado_anterior:
                    user_name = (g.user.get("name") or g.user.get("email") or "Usuario desconocido").strip()
                    comentario = (
                        f"El usuario: {user_name} ha cambiado el estado de: "
                        f"'{estado_anterior}' a '{payload.get('estado')}'."
                    )
                    usuario_id = int(g.user["id"]) if g.user.get("id") else None
                    cur.execute(
                        insert_actividad_query,
                        (consolidada_id, comentario, usuario_id),
                    )

                conn.commit()

        return jsonify({"id": row["id"], "estado": row["estado"]})

    @bank_accounts_bp.route("/cuenta-consolidada/<int:consolidada_id>/adjuntos", methods=["POST"])
    @token_required(required_roles=CUENTA_CONSOLIDADA_ROLES)
    def create_cuenta_consolidada_adjunto(consolidada_id: int):
        payload = request.get_json(silent=True) or {}
        nombre_archivo = (payload.get("nombre_archivo") or "").strip()
        mime_type = (payload.get("mime_type") or "").strip()
        archivo, decode_error = _decode_archivo_blob(payload.get("archivo"))

        if not nombre_archivo:
            return _error_response(400, "NOMBRE_ARCHIVO_REQUERIDO", "nombre_archivo is required")
        if not mime_type:
            return _error_response(400, "MIME_TYPE_REQUERIDO", "mime_type is required")
        if decode_error:
            return _error_response(400, "ARCHIVO_INVALIDO", decode_error)

        usuario_id = int(g.user["id"]) if g.user.get("id") else None

        query = """
            INSERT INTO cuenta_consolidada_adjuntos (
                cuenta_consolidada_id,
                nombre_archivo,
                mime_type,
                archivo,
                uploaded_by_user_id
            )
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id, cuenta_consolidada_id, nombre_archivo, mime_type, fecha_carga;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute("SELECT id FROM cuenta_consolidada WHERE id = %s;", (consolidada_id,))
                if cur.fetchone() is None:
                    return _error_response(404, "CONSOLIDADO_NO_ENCONTRADO", "Registro consolidado no encontrado")
                cur.execute(query, (consolidada_id, nombre_archivo, mime_type, archivo, usuario_id))
                row = cur.fetchone()
                conn.commit()

        return (
            jsonify(
                {
                    "id": row["id"],
                    "cuenta_consolidada_id": row["cuenta_consolidada_id"],
                    "nombre_archivo": row["nombre_archivo"],
                    "mime_type": row["mime_type"],
                    "fecha_carga": row["fecha_carga"].isoformat() if row["fecha_carga"] else None,
                }
            ),
            201,
        )

    @bank_accounts_bp.route("/cuenta-consolidada/<int:consolidada_id>/adjuntos", methods=["GET"])
    @token_required(required_roles=CUENTA_CONSOLIDADA_ROLES)
    def list_cuenta_consolidada_adjuntos(consolidada_id: int):
        query = """
            SELECT id, cuenta_consolidada_id, nombre_archivo, mime_type, uploaded_by_user_id, fecha_carga
            FROM cuenta_consolidada_adjuntos
            WHERE cuenta_consolidada_id = %s
            ORDER BY fecha_carga ASC, id ASC;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(query, (consolidada_id,))
                rows = cur.fetchall()

        return jsonify(
            {
                "items": [
                    {
                        "id": row["id"],
                        "cuenta_consolidada_id": row["cuenta_consolidada_id"],
                        "nombre_archivo": row["nombre_archivo"],
                        "mime_type": row["mime_type"],
                        "uploaded_by_user_id": row["uploaded_by_user_id"],
                        "fecha_carga": row["fecha_carga"].isoformat() if row["fecha_carga"] else None,
                    }
                    for row in rows
                ]
            }
        )

    @bank_accounts_bp.route("/cuenta-consolidada/<int:consolidada_id>/actividad", methods=["POST"])
    @token_required(required_roles=CUENTA_CONSOLIDADA_ROLES)
    def create_cuenta_consolidada_actividad(consolidada_id: int):
        payload = request.get_json(silent=True) or {}
        tipo = (payload.get("tipo") or "").strip()
        fecha_raw = (payload.get("fecha") or "").strip()
        comentario = payload.get("comentario")
        usuario_id = payload.get("usuario_id")

        if tipo not in ACTIVIDAD_TIPOS:
            return _error_response(400, "TIPO_ACTIVIDAD_INVALIDO", "tipo is invalid")
        if not fecha_raw:
            return _error_response(400, "FECHA_REQUERIDA", "fecha is required")
        try:
            fecha = datetime.fromisoformat(fecha_raw.replace("Z", "+00:00"))
        except ValueError:
            return _error_response(400, "FECHA_INVALIDA", "fecha must be a valid ISO datetime")

        if usuario_id is None:
            usuario_id = int(g.user["id"]) if g.user.get("id") else None
        if usuario_id is None:
            return _error_response(400, "USUARIO_ID_REQUERIDO", "usuario_id is required")

        if comentario is None:
            comentario = ""

        query = """
            INSERT INTO cuenta_consolidada_actividad (
                cuenta_consolidada_id,
                tipo,
                fecha,
                comentario,
                usuario_id
            )
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id, cuenta_consolidada_id, tipo, fecha, comentario, usuario_id;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute("SELECT id FROM cuenta_consolidada WHERE id = %s;", (consolidada_id,))
                if cur.fetchone() is None:
                    return _error_response(404, "CONSOLIDADO_NO_ENCONTRADO", "Registro consolidado no encontrado")
                cur.execute(query, (consolidada_id, tipo, fecha, comentario, usuario_id))
                row = cur.fetchone()
                conn.commit()

        return (
            jsonify(
                {
                    "id": row["id"],
                    "cuenta_consolidada_id": row["cuenta_consolidada_id"],
                    "tipo": row["tipo"],
                    "fecha": row["fecha"].isoformat() if row["fecha"] else None,
                    "comentario": row["comentario"],
                    "usuario_id": row["usuario_id"],
                }
            ),
            201,
        )

    @bank_accounts_bp.route("/cuenta-consolidada/<int:consolidada_id>/actividad", methods=["GET"])
    @token_required(required_roles=CUENTA_CONSOLIDADA_ROLES)
    def list_cuenta_consolidada_actividad(consolidada_id: int):
        query = """
            SELECT id, cuenta_consolidada_id, tipo, fecha, comentario, usuario_id
            FROM cuenta_consolidada_actividad
            WHERE cuenta_consolidada_id = %s
            ORDER BY fecha ASC, id ASC;
        """

        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=DictCursor) as cur:
                cur.execute(query, (consolidada_id,))
                rows = cur.fetchall()

        return jsonify(
            {
                "items": [
                    {
                        "id": row["id"],
                        "cuenta_consolidada_id": row["cuenta_consolidada_id"],
                        "tipo": row["tipo"],
                        "fecha": row["fecha"].isoformat() if row["fecha"] else None,
                        "comentario": row["comentario"],
                        "usuario_id": row["usuario_id"],
                    }
                    for row in rows
                ]
            }
        )

    return bank_accounts_bp
